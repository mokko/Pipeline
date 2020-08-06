""" mdl = mdvos_list
m+classic doesn't export Ausstellung as it does Sektion. It's a bug, not a feature. 
So let's a python script to exported data to write an excel list

1. select records: only records with a Ausstellung that starts with HUFO-
2. for each such record display a few fields in excel
3. objId, identNr, sachbegriff, titel, HFAusstellung, Sektion

Meant to be quick and dirty code that can grow in the future

USAGE (CLI)
mdl -i file.xml -o file.xlsx

USAGE (CLASS)
m=Mdl(input_fn, output_fn)
"""
import os 
import sys
from openpyxl import Workbook, load_workbook
from lxml import etree

class Mdl:
    def __init__(self, input_fn, output_fn): 

        self._prepare_wb (output_fn)
        self.ns = {'m':'http://www.mpx.org/mpx'}

        tree = etree.parse (input_fn)
        #only records which have an exhibit in HUFO
        r = tree.xpath("/m:museumPlusExport/m:sammlungsobjekt[starts-with(m:ausstellung, 'HUFO -')]", 
            namespaces = self.ns)

        for so_node in r:
            self.add_row (so_node)

        self.wb.save (filename = output_fn)

    def _prepare_wb (self, output_fn):
        if os.path.isfile (output_fn):
            print(f"Warning: File {output_fn} exists already, will be overwritten!")
        self.wb = Workbook()
        ws1 = self.wb.active
        ws1.title = "MDVOS Liste"
        ws1['A1']="objId"
        ws1['B1']="IdentNr"
        ws1['C1']="Sachbegriff"
        ws1['D1']="Titel"
        ws1['E1']="Ausstellung [Sektion]"
        ws1.column_dimensions['C'].width = 30
        ws1.column_dimensions['D'].width = 45
        ws1.column_dimensions['E'].width = 90

    def _xpathify (self, node, xpath):
        """apply xpath to node and stringify result"""
        
        r = node.xpath (xpath, namespaces = self.ns)
        if isinstance(r, list):
            return '; '.join(r)
        else:
            return r

    def _ausify (self, node):
        """apply xpath for Ausstellung and return string"""
        
        r = node.xpath ("m:ausstellung", namespaces = self.ns)
        my_list=[]
        if isinstance (r, list):
            for each in r:
                sektion=each.get('sektion')
                #show only HUFO exhibits
                if each.text.startswith('HUFO -'):
                    my_list.append (f"{each.text} [{sektion}]")
        else:
            print('Error: Should always be a list!')
            sys.exit(1)
        return '; '.join (my_list)

    def add_row (self,so_node):
        #3. objId, identNr, sachbegriff, titel, HFAusstellung/Sektion
        #text() can return list
        ws1 = self.wb.active
        new_row = str (ws1.max_row+1)
        print (new_row)
        ws1['A'+new_row] = so_node.get ("objId") # there can be only one
        ws1['B'+new_row] = self._xpathify (so_node,"m:identNr/text()")
        ws1['C'+new_row] = self._xpathify (so_node, "m:sachbegriff/text()")
        ws1['D'+new_row] = self._xpathify (so_node, "m:titel/text()")
        ws1['E'+new_row] = self._ausify (so_node)

        #print (f"{objId}|{identNr}|{sb}|{t}|{aus}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument ('-i', '--input')
    parser.add_argument ('-o', '--output')

    args = parser.parse_args()
    m = Mdl (args.input, args.output)
