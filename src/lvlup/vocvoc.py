"""From several smaller xlsx translation tables to one bigger xml dictionary.

USAGE:
    vocvoc(src_dir, out_xml)

assembles translations in xml to make them available to xslt looking 
recursively through src_dir.

FORMAT IS ROUGHLY LIKE THIS:
    <mpxvok>
        <context name="titel@art">
            <concept> <!--no id-->
                <pref lang="de" src="EM-SM>Titel</pref>
                <pref lang="en">Title</pref>
                <syn>Titellei</syn>
            </concept>
        </context>
    </mpxvok>

Let's not write any term with frequency = 0 to the vocabulary. vokvok
overwrites the result file every time it's called.

loops thru all sheets of translation.xlsx

We cannot merge them up to this point, because it would mess up the frequency
count. So I don't just write one big translation file to begin with.

Second question is if we can merge translations from different elements/fields.
I guess that would be just about possible. What are the chances that two 
different fields have the same term, but insist on translating them 
differently? Not so big. I mean "London" in one field will still be "London" in
the next field. Okay, but to build a structure where that is not possible even
a single time? I don't think so. Hence the terms in format are specific to 
contexts.

PROBLEMS
It's still possible to have multiple translations of the same term. At the 
moment, they should all be called pref[@lang="en"]. That's not acceptable.
"""

import os
import glob
import openpyxl
from lxml import etree as ET
from openpyxl import Workbook, load_workbook

needle_fn = "translate.xlsx"


class vocvoc:
    def __init__(self, in_fn):
        if os.path.exists(in_fn):
            self.in_fn = in_fn
        else:
            raise FileNotFoundError("vocvoc: Input file/dir not found!")
        print("*vocvoc")

    def recursive(self, out_fn):
        print(f"**vocvoc source dir {self.in_fn}")
        root = ET.Element("mpxvoc")  # start a new document
        needle_path = os.path.realpath(os.path.join(self.in_fn, f"./**/{needle_fn}"))
        for path in glob.iglob(needle_path, recursive=True):
            wb = self._prepare_wb(path)
            print(f"*Processing translation table: {path}")
            for sheet in wb.worksheets:
                print(f"   {sheet.title}")
                self._per_sheet(sheet, root, path)
        self._write_xml(root, out_fn)

    def single(self, out_fn):
        print(f"**vocvoc in_fn {self.in_fn}")
        root = ET.Element("mpxvoc")  # start a new document
        wb = self._prepare_wb(self.in_fn)
        for sheet in wb.worksheets:
            print(f"   {sheet.title}")
            self._per_sheet(sheet, root, self.in_fn)
        self._write_xml(root, out_fn)

    #
    ### PRIVATE ###
    #
    def _add_concept(self, xml, context, row, scope):
        # change in place is not Pythonic
        # i just want value and strip, why is this difficult?
        new = list()
        for n in range(4):
            try:
                value = row[n].value
            except:
                value = None
            if type(value) is str:
                value = value.strip().replace('"', r"'")
            new.append(value)

        term_xls = new[0]
        translation_xls = new[1]
        comment_xls = new[2]
        freq_xls = new[3]  # xml attribs must be string
        try:
            src = new[4]
        except:
            src = None

        # print (f"context:{context}")
        # (1)Does concept exist already?
        rls = xml.xpath(f'//mpxvoc/context[@name = "{context}"]')
        if len(rls) > 0:
            context_nd = rls[0]
        else:
            # print (f"\tContext doesn't exists yet: {context}")
            context_nd = ET.SubElement(xml, "context", attrib={"name": context})
        # (2)Does pref_de exist yet?
        rls = context_nd.xpath(f"./concept/pref[@lang='de' and .=\"{term_xls}\"]")
        if len(rls) > 0:
            pref_de = rls[0]
            concept_nd = pref_de.xpath(f"..")[0]
            freq_xml = int(concept_nd.get("freq"))
            concept_nd.set("freq", str(freq_xml + freq_xls))
        else:
            # print (f"\tconcept/pref doesn't exist yet: {term_xls} ")
            concept_nd = ET.SubElement(
                context_nd, "concept", attrib={"freq": str(freq_xls)}
            )
            pref_de = ET.SubElement(concept_nd, "pref", attrib={"lang": "de"})
            pref_de.text = term_xls
        # (3)Does translation exist yet?

        xpath = f"../pref[@lang = 'en' and . = \"{translation_xls}\"]"
        rls = pref_de.xpath(xpath)
        if len(rls) > 0:
            pref_en = rls[0]
        else:
            # print (f"\ttranslation doesn't exist yet {translation}")
            pref_en = ET.SubElement(concept_nd, "pref", attrib={"lang": "en"})
            pref_en.text = translation_xls
        # there should always be scope
        scope_nd = ET.SubElement(concept_nd, "scope")
        scope_nd.text = scope
        if comment_xls:
            comment_nd = ET.SubElement(concept_nd, "comment")
            comment_nd.text = comment_xls
        if src is not None:
            # just append another element sources if multiple
            sources_nd = ET.SubElement(concept_nd, "sources")
            sources_nd.text = src
        # https://stackoverflow.com/questions/40154757/sorting-xml-tags-by-child-elements-python
        concept_nd[:] = sorted(concept_nd, key=lambda e: e.tag)

    def _mk_scope(self, path):
        npath = os.path.dirname(os.path.abspath(path))
        return npath.replace("\\", "/")

    def _per_sheet(self, sheet, xml, path):
        lno = 1  # 1-based line counter
        for term_xls in sheet["A"]:
            if lno > 1:  # IGNORE HEADER
                scope = self._mk_scope(path)
                translation = sheet[f"B{lno}"].value
                freq = int(sheet[f"D{lno}"].value)
                if freq > 0 and translation is not None:
                    self._add_concept(xml, sheet.title, sheet[lno], scope)
            lno += 1

    def _prepare_wb(self, xls_fn):
        """Read existing xls or make new one.

        Returns workbook."""

        if os.path.isfile(xls_fn):
            # print (f'   Excel file exists ({xls_fn})')
            return load_workbook(filename=xls_fn)
        else:
            print(f"   Excel file doesn't exist yet, making it ({xls_fn})")
            return Workbook()

    def _write_xml(self, root, out_fn):
        ET.indent(root)
        doc = ET.ElementTree(root)
        out_fn = os.path.realpath(out_fn)
        print(f"**About to write to {out_fn}, overwriting old file")
        with open(out_fn, "wb") as f:
            doc.write(
                f,
                encoding="UTF-8",
                method="xml",
                xml_declaration=True,
                pretty_print=True,
            )


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Manually create a mpxvoc.xml file from a translate.xlsx"
    )
    parser.add_argument(
        "-i",
        "--input",
        help="specificy location of translate.xlsx",
        required=True,
    )
    args = parser.parse_args()
    args.output = "mpxvoc.xml"  # quick and dirty default

    # execute from usual dir data/scope/date
    t = vocvoc(args.input)
    t.single(args.output)
